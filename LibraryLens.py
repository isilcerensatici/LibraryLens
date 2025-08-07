from pyzbar.pyzbar import decode
import cv2
import requests
import openpyxl
import os
import time

# Fetch book information from Google Books API
def fetch_book_info(isbn):
    url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
    try:
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        data = response.json()

        if "items" in data:
            book = data["items"][0]["volumeInfo"]
            title = book.get("title", "").strip()
            authors = book.get("authors", [])

            if title and authors:
                return title, ", ".join(authors)
    except requests.exceptions.RequestException as e:
        print(f"API connection error: {e}")
    
    return None, None

# Create Excel file if it doesn't exist
def prepare_excel(file_name):
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["ISBN", "Title", "Author"])
        workbook.save(file_name)

# Save data to Excel
def save_to_excel(file_name, isbn, title, author):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    sheet.append([isbn, title, author])
    workbook.save(file_name)

# Read previously saved ISBNs from Excel
def read_existing_isbns(file_name):
    existing_isbns = set()
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            isbn = str(row[0]).strip()
            existing_isbns.add(isbn)
    return existing_isbns

# Barcode scanning process
def scan_barcode():
    cap = cv2.VideoCapture(1)
    excel_file = "books.xlsx"
    prepare_excel(excel_file)

    scanned_isbns = read_existing_isbns(excel_file)

    show_message = False
    message_time = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        barcodes = decode(frame)
        for barcode in barcodes:
            isbn = barcode.data.decode("utf-8").strip()

            if isbn not in scanned_isbns:
                title, author = fetch_book_info(isbn)

                if title and author:
                    scanned_isbns.add(isbn)
                    print(f"Barcode Scanned: {isbn}")
                    print(f"Title: {title}, Author: {author}")

                    save_to_excel(excel_file, isbn, title, author)

                    show_message = True
                    message_time = time.time()
                else:
                    print(f"No book info found for: {isbn}")

            (x, y, w, h) = barcode.rect
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        # Display message for 2 seconds
        if show_message and time.time() - message_time < 2:
            # Shadow effect (black)
            cv2.putText(frame, "Saved", (50, 50),
                        cv2.FONT_HERSHEY_DUPLEX, 1.2, (0, 0, 0), 4)
            # Main text (white)
            cv2.putText(frame, "Saved", (50, 50),
                        cv2.FONT_HERSHEY_DUPLEX, 1.2, (255, 255, 255), 2)
        else:
            show_message = False

        cv2.imshow("Barcode Scanner", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# Start the program
scan_barcode()